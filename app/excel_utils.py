"""
Экспорт списка команд в Excel
Ничего менять не нужно
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import io

def export_to_excel(teams):
    """
    Создаёт Excel-файл со списком всех команд
    Возвращает байты файла
    """
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Команды мероприятия"
    
    # Заголовки колонок
    headers = ["ID", "Название команды", "Капитан", "Email капитана", 
               "Кол-во участников", "Учебное заведение", "Дата регистрации"]
    
    # Стиль для заголовков
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Записываем заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
    
    # Записываем данные
    for row, team in enumerate(teams, 2):
        ws.cell(row=row, column=1, value=team.id)
        ws.cell(row=row, column=2, value=team.team_name)
        ws.cell(row=row, column=3, value=team.captain_name)
        ws.cell(row=row, column=4, value=team.captain_email)
        ws.cell(row=row, column=5, value=team.participant_count)
        ws.cell(row=row, column=6, value=team.institution or "—")
        ws.cell(row=row, column=7, value=team.created_at.strftime("%d.%m.%Y %H:%M") if team.created_at else "—")
    
    # Автоматическая ширина колонок
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = adjusted_width
    
    # Сохраняем в байтовый поток
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()